import datetime
from importlib.abc import MetaPathFinder
import json
from math import inf
from sys import path
from openpyxl.styles.fills import PatternFill
import pytz
import requests
import shutil
import codecs
import os
from dotenv import load_dotenv
from PyPDF2 import PdfFileMerger
from reportlab.pdfgen.canvas import Canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.units import inch
import logging
from logging.handlers import RotatingFileHandler
import openpyxl
from dateutil import tz
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Protection, Font
from openpyxl.styles import Alignment
from openpyxl.formatting.rule import FormulaRule
from reportlab.graphics import renderPDF
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from svglib.svglib import svg2rlg


pdfmetrics.registerFont(TTFont('FreeSans', 'fonts/FreeSans.ttf'))
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
log_dir = os.path.join(BASE_DIR, 'logs/')
log_file = os.path.join(BASE_DIR, 'logs/stickers.log')
console_handler = logging.StreamHandler()
file_handler = RotatingFileHandler(
    log_file,
    maxBytes=100000,
    backupCount=3,
    encoding='utf-8'
)
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s, %(levelname)s, %(message)s',
    handlers=(
        file_handler,
        console_handler
    )
)


dotenv_path = os.path.join(os.path.dirname(__file__), '.env')
if os.path.exists(dotenv_path):
    load_dotenv(dotenv_path)


def create_path_if_not_exist(path):
    if not os.path.exists(path):
        os.mkdir(path)


def get_now_time():
    d = datetime.datetime.utcnow()
    d_with_timezone = d.replace(tzinfo=pytz.UTC)
    return(d_with_timezone.isoformat())


TOKEN = os.environ.get('TOKEN')
base_url_for_getting_orders = 'https://suppliers-api.wildberries.ru/api/v2/orders'
headers = {
    'Authorization': TOKEN,
}


cred = json.load(open('credentials.json', 'rb'))
TOKEN = cred['БелотеловАГ']['token']
NAME = cred['БелотеловАГ']['name']
headers = {
    'Authorization': TOKEN,
}


def create_all_today_path():
    pdf_path = os.path.join(BASE_DIR, 'pdf/')
    create_path_if_not_exist(pdf_path)
    today_prefix_path = os.path.join(
        pdf_path, datetime.datetime.today().strftime('%Y_%m_%d'))
    create_path_if_not_exist(today_prefix_path)
    today_path_with_name = os.path.join(today_prefix_path, NAME)
    create_path_if_not_exist(today_path_with_name)
    backup_dir = os.path.join(today_path_with_name, 'results/')
    json_dir = os.path.join(today_path_with_name, 'json/')
    create_path_if_not_exist(backup_dir)
    create_path_if_not_exist(json_dir)
    return {
        'backup_dir': backup_dir,
        'json_dir': json_dir,
        'today_prefix_path': today_prefix_path,
        'today_path_with_name': today_path_with_name
    }


MEDIUM_BORDER = Border(left=Side(style='medium'),
                       right=Side(style='medium'),
                       top=Side(style='medium'),
                       bottom=Side(style='medium'))
THIN_BORDER = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))


def create_pdf_stickers_by_ids(ids):
    url_for_getting_stikers = 'https://suppliers-api.wildberries.ru/api/v2/orders/stickers/pdf'
    data_for_pdf = ''
    pdfs = []
    for id in ids:
        json_orders_id = {
            "orderIds": [int(id)]
        }
        response = requests.post(
            url_for_getting_stikers,
            json=json_orders_id,
            headers=headers)
        data_for_pdf = response.json()['data']['file']
        file_data = bytes(data_for_pdf, 'utf-8')
        today_path_with_name = create_all_today_path()['today_path_with_name']
        path = os.path.join(today_path_with_name, f'{id}.pdf')
        with open(path, 'wb') as f:
            f.write(codecs.decode(file_data, 'base64'))
        pdfs += [path]
    merger = PdfFileMerger()
    for pdf in pdfs:
        merger.append(pdf)
    merger.write("result2.pdf")
    merger.close()


def get_all_orders(status=0, date_end=get_now_time(),
                   date_start='2020-04-05T00:47:17.528082+00:00'):
    logging.info(f'Получение всех заказов со статусом {status}')
    date_end = get_now_time()
    orders = []
    params = {
        'date_end': date_end,
        'date_start': date_start,
        'status': status,
        'take': 1000,
        'skip': 0
    }
    response = requests.get(
        base_url_for_getting_orders,
        headers=headers,
        params=params)
    try:
        orders_from_current_response = response.json()['orders']
    except KeyError as e:
        orders_from_current_response = []
        logging.error(e, exc_info=True)
    orders += orders_from_current_response
    while orders_from_current_response != []:
        params['skip'] += len(orders_from_current_response)
        response = requests.get(
            base_url_for_getting_orders,
            headers=headers,
            params=params)
        orders_from_current_response = response.json()['orders']
        orders += orders_from_current_response
        logging.info(f'{len(orders)}')
    logging.info(f'Получено {len(orders)}')
    orders = clean_orders_from_user_status_1(orders)
    return orders


def clean_orders_from_user_status_1(orders):
    logging.info('Очистка заказов от заказов со статусом 1')
    filtered_orders = []
    for order in orders:
        if order['userStatus'] != 1:
            filtered_orders += [order]
    logging.info(f'Осталось  {len(filtered_orders)}')
    return filtered_orders


def barcode_key_for_sorting(order):
    return int(order['barcode'])


def get_orders_ids(orders):
    ids = []
    for order in orders:
        ids.append(int(order['orderId']))
    return ids


def get_barcodes_with_orders_and_chartId(orders):
    barcodes_and_ids = {}
    logging.info('Сортировка информации по баркодам')
    for order in orders:
        barcode = order['barcode']
        id = int(order['orderId'])
        chrt_id = order['chrtId']
        if barcode not in barcodes_and_ids.keys():
            barcodes_and_ids[barcode] = {'orders': [id], 'chrtId': chrt_id}
        else:
            barcodes_and_ids[barcode]['orders'] += [id]

    logging.info(f'Получено {len(barcodes_and_ids)} баркодов')
    return barcodes_and_ids


def edit_blank_pdf(barcode_info):

    canvas = Canvas('pdf/blank1.pdf', pagesize=(1.6 * inch, 1.2 * inch))
    font_size = 9
    canvas.setFont('FreeSans', font_size)
    n = 0
    blanks_number = 1
    slice = 23
    name_of_host = NAME
    name = barcode_info['name']
    size = barcode_info['size']
    article = f'art {barcode_info["article"]}'
    color = barcode_info['color']
    extra_colors = barcode_info['extra_colors']
    count = 'Количество = ' + str(barcode_info['count'])
    chrtId = f'chrtId = {barcode_info["chrtId"]}'
    blanks = ['pdf/blank1.pdf']
    date = datetime.datetime.now().strftime("%d.%m.%y  %H:%M")
    for params in (count, name_of_host, date, name, size,
                   color, extra_colors, article, chrtId):
        a = str(params)
        while len(a) > 0:
            if n > 8:
                blanks_number += 1
                canvas.save()
                canvas = Canvas(
                    f'pdf/blank{blanks_number}.pdf', pagesize=(1.6 * inch, 1.2 * inch))
                canvas.setFont('FreeSans', font_size)
                blanks += [f'pdf/blank{blanks_number}.pdf']
                n = 0
            canvas.drawString(2, 75 - n * font_size,
                              a[:slice], charSpace=0)
            a = a[slice:]
            n += 1
    canvas.save()
    merg = PdfFileMerger()
    for blank in blanks:
        merg.append(blank)
    merg.write('pdf/blank.pdf')
    merg.close()
    logging.info(
        f'Получена информация: наименование "{barcode_info["name"]}" '
        f' {count}, '
        f'артикул = {article}, '
        f'chrtId = {chrtId} '
        f'{"Size "+size if size != "" else ""} '
        f'{"color "+ color if color != "" else ""} '
        f'{"Доп цвета: "+extra_colors if extra_colors != "" else ""} ')


def get_StickerEncoded_by_orderId(id):
    url = 'https://suppliers-api.wildberries.ru/api/v2/orders/stickers'
    json_order_id = {
        "orderIds": [int(id)]
    }
    response = requests.post(
        url,
        json=json_order_id,
        headers=headers)
    return response.json()['data'][0]['sticker']['wbStickerEncoded']


def create_and_merge_pdf_by_barcodes_and_ids(barcodes_and_ids):

    logging.info('Создание pdf для баркодов')
    results_files = []
    url_for_getting_stikers = 'https://suppliers-api.wildberries.ru/api/v2/orders/stickers/pdf'
    for barcode in barcodes_and_ids.keys():
        edit_blank_pdf(barcodes_and_ids[barcode]['info'])
        pdfs = ['pdf/blank.pdf']
        # for id in barcodes_and_ids[barcode]['orders']:
        #     json_orders_id = {
        #         "orderIds": [int(id)]
        #     }
        #     response = requests.post(
        #         url_for_getting_stikers,
        #         json=json_orders_id,
        #         headers=headers)
        #     data_for_pdf = response.json()['data']['file']
        #     file_data = bytes(data_for_pdf, 'utf-8')
        #     today_path_with_name = create_all_today_path()['today_path_with_name']
        #     path = os.path.join(today_path_with_name, f'{id}.pdf')
        #     with open(path, 'wb') as f:
        #         f.write(codecs.decode(file_data, 'base64'))
        #     pdfs += [path]
        merger = PdfFileMerger()
        orders = barcodes_and_ids[barcode]['orders']
        json_orders_id = {
            "orderIds": orders
        }
        start = 0
        end = 1000
        while orders[start:end] != []:
            json_orders_id = {
                "orderIds": orders[start:end]
            }
            response = requests.post(
                url_for_getting_stikers,
                json=json_orders_id,
                headers=headers)
            data_for_pdf = response.json()['data']['file']
            while len(data_for_pdf) < 1105:
                response = requests.post(
                    url_for_getting_stikers,
                    json=json_orders_id,
                    headers=headers)
                data_for_pdf = response.json()['data']['file']

            file_data = bytes(data_for_pdf, 'utf-8')
            today_path_with_name = create_all_today_path()[
                'today_path_with_name']
            path = os.path.join(today_path_with_name, f'{barcode}_{end}.pdf')
            with open(path, 'wb') as f:
                f.write(codecs.decode(file_data, 'base64'))
            start += 1000
            end += 1000
            pdfs += [path]

        # response = requests.post(
        #         url_for_getting_stikers,
        #         json=json_orders_id,
        #         headers=headers)
        # data_for_pdf = response.json()['data']['file']
        # while len(data_for_pdf) <1105:
        #         response = requests.post(
        #         url_for_getting_stikers,
        #         json=json_orders_id,
        #         headers=headers)
        #         data_for_pdf = response.json()['data']['file']
        # file_data = bytes(data_for_pdf, 'utf-8')
        # today_path_with_name = create_all_today_path()['today_path_with_name']
        # path = os.path.join(today_path_with_name, f'{barcode}.pdf')
        # with open(path, 'wb') as f:
        #     f.write(codecs.decode(file_data, 'base64'))
        # pdfs += [path]

        for pdf in pdfs:
            merger.append(pdf)
        path_for_result_of_barcode = os.path.join(
            today_path_with_name, f'result_{barcode}.pdf')
        merger.write(path_for_result_of_barcode)
        results_files.append(path_for_result_of_barcode)
        merger.close()
        logging.info(f'Создано pdf для баркода {barcode}')
    return results_files


def create_pdf_stickers_by_barcodes(barcodes_and_ids):
    results_files = []
    results_files = create_and_merge_pdf_by_barcodes_and_ids(barcodes_and_ids)
    merger = PdfFileMerger()
    logging.info('Объединение pdf файлов в results.pdf')
    for result in results_files:
        merger.append(result)
    merger.write('results.pdf')
    merger.close()
    add_results_file_to_today_backup('results.pdf')


def get_data_nomenclature_from_card_by_chrtId(card, chrtId):
    all_nomenclatures = card['nomenclatures']
    for nomenclature in all_nomenclatures:
        for diffetent_types in nomenclature['variations']:
            vendorCode = nomenclature['vendorCode']
            data_about_nomenclature = diffetent_types
            for field in data_about_nomenclature:
                if data_about_nomenclature[field] == chrtId:
                    return(data_about_nomenclature, vendorCode, nomenclature)


def get_card_by_chrtId(chrtId):
    url = 'https://suppliers-api.wildberries.ru/card/list'
    json_for_request = {
        "id": 1,
        "jsonrpc": "2.0",
        "params": {
            "filter": {
                "find": [
                    {
                        "column": "nomenclatures.variations.chrtId",
                        "search": chrtId
                    }
                ],
                "order": {
                    "column": "string",
                    "order": "string"
                }
            }
        }
    }
    response = requests.post(url=url, headers=headers, json=json_for_request)
    card = response.json()['result']['cards'][0]
    return card


def get_card_by_nmid(nmid):
    url = 'https://suppliers-api.wildberries.ru/card/list'
    json_for_request = {
        "id": 1,
        "jsonrpc": "2.0",
        "params": {
            "filter": {
                "find": [
                    {
                        "column": "nomenclatures.nmId",
                        "search": nmid
                    }
                ],
                "order": {
                    "column": "string",
                    "order": "string"
                }
            }
        }
    }
    response = requests.post(url=url, headers=headers, json=json_for_request)
    card = response.json()['result']['cards'][0]
    return card


def getting_information_about_barcode_by_chartId(chrtId):
    print(1)
    good = get_card_by_chrtId(chrtId)
    supplierVendorCode = good['supplierVendorCode']
    name = ''
    nomenclature_data, article, nomenclature = get_data_nomenclature_from_card_by_chrtId(
        good, int(chrtId))
    article = supplierVendorCode + article
    nmid = nomenclature['nmId']
    size = ''
    color = ''
    extra_colors = ''
    addin = nomenclature['addin']
    for type_and_params in addin:
        if type_and_params['type'] in ('Цвет', 'Основной цвет'):
            color = type_and_params['params'][0]['value']
        if type_and_params['type'] == 'Доп. цвета':
            for extra_color in type_and_params['params']:
                extra_colors += extra_color['value'] + ' '
    if 'Доп. цвета' in nomenclature.keys():
        extra_colors = nomenclature['Доп. цвета']
    for type_and_params in good['addin']:
        if type_and_params['type'] == 'Наименование':
            name = type_and_params['params'][0]['value']
    for data in nomenclature_data['addin']:
        if data['type'] == 'Размер':
            size = data['params'][0]['value']
        if data['type'] == 'Доп. цвета':
            extra_colors = data['params'][0]['value']
    info = {
        'name': name,
        'article': article,
        'chrtId': chrtId,
        'size': size,
        'color': color,
        'extra_colors': extra_colors,
        'nmid': nmid
    }
    return info


def add_information_about_barcodes_and_len(barcodes):
    for barcode in barcodes.keys():
        barcodes[barcode]['info'] = getting_information_about_barcode_by_chartId(
            barcodes[barcode]['chrtId'])
        barcodes[barcode]['info']['count'] = len(barcodes[barcode]['orders'])
    return barcodes


def input_colum_names(column_names, sheet, border=MEDIUM_BORDER):
    column = 1
    row = 1
    for column_name in column_names:
        cell = sheet.cell(row=row, column=column)
        cell.value = column_name
        cell.border = border
        column += 1

def input_row(values, sheet, row, start_column=1, border=MEDIUM_BORDER):
    column = start_column
    for value in values:
        cell = sheet.cell(row=row, column=column)
        cell.value = value
        cell.border = border
        column += 1

def create_db_for_checking(barcodes):
    barcodes_and_stickers = {}
    logging.info('Получение расшифрованных стикеров')
    article_counts = {}
    for barcode in barcodes.keys():
        logging.info(f'Получение расшифрованных стикеров для {barcode}')
        barcodes_and_stickers[barcode] = {}
        arcticle = barcodes[barcode]['info']['article']
        name = barcodes[barcode]['info']['name']
        size = barcodes[barcode]['info']['size']
        color = barcodes[barcode]['info']['color']
        orders_and_sticker_encoded = get_orderId_and_sticker_encoded(
            barcodes[barcode]['orders'])
        for order in barcodes[barcode]['orders']:
            sticker_encoded = orders_and_sticker_encoded[order]
            barcodes_and_stickers[barcode][order] = {
                'sticker_encoded': sticker_encoded,
                'article': arcticle,
                'name': name[:40],
                'size': size,
                'color': color
            }
            if not arcticle in article_counts:
                article_counts[arcticle] = 1
            else:
                article_counts[arcticle] += 1

    book = openpyxl.Workbook()
    sheet = book.active
    book.create_sheet("Итог")
    book.create_sheet('Проверка')
    sheet.protection.sheet = True
    sheet['A1'] = 'Номер заказа'
    sheet['B1'] = 'Артикул'
    sheet['C1'] = 'Наименование'
    sheet['D1'] = 'Баркод'
    sheet['E1'] = 'Stick'
    sheet['F1'] = 'Размер'
    sheet['G1'] = 'Цвет'
    sheet['H1'] = 'Проверено?'

    column_names = ('Номер заказа', 'Артикул', 'Наименование', 'Баркод', 'Stick', 'Размер', 'Цвет', 'Проверено?')
    input_colum_names(column_names, sheet)

    sheet.protection.sheet = True
    sheet.protection.set_password('werock')
    sheet.protection.enable()
    green_fill = PatternFill(bgColor="a0db8e")
    sheet.conditional_formatting.add('F2:F3000', FormulaRule(
        formula=['IF(F2="Да",True,False)'], stopIfTrue=True, fill=green_fill))
    row = 2
    logging.info('Формирование xlsx файла')
    sheet.column_dimensions['A'].width = 10
    sheet.column_dimensions['D'].width = 15
    sheet.column_dimensions['E'].width = 10
    for barcode in barcodes_and_stickers.keys():
        barcode_info = barcodes_and_stickers[barcode]
        for order in barcode_info.keys():
            info = barcode_info[order]
            article = info['article']
            sticker_encoded = info['sticker_encoded']
            name = info['name']
            values = (str(order), str(article), name, str(barcode), sticker_encoded, info['size'], info['color'], f'=IF(NOT(ISERROR(MATCH(E{row},Проверка!A:A,0))),"Да","")' )
            input_row(values, sheet, row, start_column=1, border=MEDIUM_BORDER)
            row += 1

    book.active = 1
    sheet = book.active
    sheet.protection.sheet = True
    column_names = ('Баркод', 'Артикул', 'Размем', 'Цвет', 'Должно быть', 'Проверено', 'Комментарий')
    input_colum_names(column_names, sheet)
    
    sheet.column_dimensions['A'].width = 15

    row = 2

    for barcode in barcodes_and_stickers.keys():
        column = 1
        info = barcodes[barcode]['info']
        count = len(barcodes[barcode]['orders'])

        values = (barcode, info['article'], info['size'], info['color'], count, f'=COUNTIF(Проверка!A:A,A{row})', '')
        input_row(values, sheet, row, start_column=1)

        row += 1


    cell = sheet.cell(row=row, column=5)
    cell.value = f'=SUM(E2:E{row-1})'
    cell.border = MEDIUM_BORDER
    cell = sheet.cell(row=row, column=6)
    cell.value = f'=SUM(F2:F{row-1})'
    cell.border = MEDIUM_BORDER

    for row in sheet.iter_rows():
        for cell in row:      
            cell.alignment = Alignment(wrap_text=True,vertical='top') 


    book.active = 2
    sheet = book.active
    red_fill = PatternFill(bgColor="FFC7CE")

    sheet.conditional_formatting.add('B1:B3000', FormulaRule(
        formula=['IF(B1="Ошибка",True,False)'], stopIfTrue=True, fill=red_fill))

    row = 2
    sheet['A1'] = 'Данные'
    sheet['B1'] = 'Результат'
    sheet['D1'] = 'Артикул'
    sheet.protection.sheet = True
    sheet.protection.enable()
    for barcode in barcodes_and_stickers.keys():
        barcode_info = barcodes_and_stickers[barcode]
        for order in barcode_info.keys():
            for i in range(2):
                if row % 2 == 0:
                    sheet[f'A{row}'].border = Border(left=Side(style='medium'), right=Side(
                        style='thin'), top=Side(style='medium'), bottom=Side(style='thin'))
                    sheet[f'B{row}'] = f'=IF(AND(ISERROR(MATCH(A{row+1},C{row},0)),A{row}<>"",A{row+1}<>""),"Ошибка","")'
                    sheet[f'B{row}'].border = Border(left=Side(style='thin'), right=Side(
                        style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))
                    sheet[f'C{row}'] = f'=IF(ISERROR(INDEX(Sheet!D:D,MATCH(A{row},Sheet!E:E,0),1)),"",INDEX(Sheet!D:D,MATCH(A{row},Sheet!E:E,0),1))'
                else:
                    sheet[f'A{row}'].border = Border(left=Side(style='thin'), right=Side(
                        style='medium'), top=Side(style='thin'), bottom=Side(style='medium'))
                    sheet[f'D{row}'] = f'=IF(AND(B{row-1}<>"Ошибка", A{row-1}<>""),INDEX(Sheet!B:B, MATCH(A{row},Sheet!D:D,0),1), "")'
                sheet[f'A{row}'].number_format = '@'
                sheet[f'A{row}'].protection = Protection(locked=False)
                row += 1
    for column in ['C', 'D']:
        for cell in sheet[column]:
            cell.font = Font(color='Ffffffff')
    book.save('db.xlsx')
    book.close()


def check_and_delete_orders_with_blank_officeAddress(orders):
    logging.info('Производится отчистка от заказов без адреса')
    count = 0
    for order in orders:
        if order['officeAddress'] == "":
            count += 1
            orders.remove(order)

    return (orders, count)


def set_status_to_orders_by_ids(status, ids):
    url_for_set_status = 'https://suppliers-api.wildberries.ru/api/v2/orders'
    data_for_bulk_set_status = []

    for id in ids:
        data_for_bulk_set_status += [{
            "orderId": str(id),
            "status": int(status)
        }]

    js = json.dumps(data_for_bulk_set_status)
    response = requests.put(url_for_set_status, headers=headers, data=js)
    logging.info(response.content.decode('utf-8'))
    print(ids)


def set_status_to_orders(status, orders):
    url_for_set_status = 'https://suppliers-api.wildberries.ru/api/v2/orders'
    data_for_bulk_set_status = []
    for order in orders:
        data_for_bulk_set_status += [{
            "orderId": order["orderId"],
            "status": int(status)
        }]
    js = json.dumps(data_for_bulk_set_status)
    logging.info(js)
    response = requests.put(url_for_set_status, headers=headers, data=js)
    logging.info(response.json())

def get_barcodes_with_full_info(orders):
    barcodes = get_barcodes_with_orders_and_chartId(orders)
    print(1)
    barcodes = add_information_about_barcodes_and_len(barcodes)

    print(2)
    barcodes = sorted_barcodes_by_count_of_orders(barcodes)
    print(3)
    return barcodes


def create_stickers():
    orders = get_all_orders(status=1)
    if len(orders) == 0:
        return (0, 0)
    barcodes = get_barcodes_with_full_info(orders)
    with open('barcodes.json', 'w', encoding='utf-8') as f:
        json.dump(barcodes, f, ensure_ascii=False)
    add_json_file_to_today_json('barcodes.json')
    create_pdf_stickers_by_barcodes(barcodes)
    return (len(orders), barcodes)

def create_stickers_by_orders(orders):
    if len(orders) == 0:
        return (0, 0)
    barcodes = get_barcodes_with_full_info(orders)
    with open('barcodes.json', 'w', encoding='utf-8') as f:
        json.dump(barcodes, f, ensure_ascii=False)
    add_json_file_to_today_json('barcodes.json')
    create_pdf_stickers_by_barcodes(barcodes)
    return (len(orders), barcodes)


def filter_orders_by_barcode(orders, barcode):
    filtered_barcodes = []
    for order in orders:
        if order['barcode'] == barcode:
            filtered_barcodes += [order]
    logging.info(
        f'Получено {len(filtered_barcodes)} заказов с баркодом = {barcode}')
    return filtered_barcodes


def sorted_barcodes_by_count_of_orders(barcodes):
    sorted_tuples = sorted(barcodes.items(), key=lambda x: len(
        x[1]['orders']), reverse=True)
    sorted_dict = {k: v for k, v in sorted_tuples}
    return sorted_dict


def set_status_collected_for_all_on_assembly():
    orders = get_all_orders(status=1)
    set_status_to_orders(2, orders)
    logging.info(f'{len(orders)} заказов переведены в собранные')
    return len(orders)


def sorted_by_barcode_set_status_on_assembly(barcode, limit):
    orders = get_all_orders(status=0)
    orders = filter_orders_by_barcode(orders, barcode)[:limit]
    set_status_to_orders(1, orders)


def set_status_on_assmebly_by_limit_and_date(limit=350):
    orders = get_all_orders(status=0)
    orders = orders[-limit:]
    set_status_to_orders(1, orders)


def set_status_on_assmebly_by_limit(limit=350):
    orders = get_all_orders(status=0)
    orders = orders[:limit]
    set_status_to_orders(1, orders)


def create_stickers_by_id(ids):
    date_end = get_now_time()
    orders = []
    for id in ids:
        date_start = '2021-11-06T00:47:17.528082+00:00'

        params = {
            'date_end': date_end,
            'date_start': date_start,
            'status': 2,
            'take': 100,
            'skip': 0,
            'id': id
        }
        response = requests.get(
            base_url_for_getting_orders,
            headers=headers,
            params=params)
        try:
            orders_from_current_response = response.json()['orders']
            print(orders_from_current_response)
        except KeyError as e:
            logging.error(e, exc_info=True)
            print(id)
        orders += orders_from_current_response
        logging.info(f'Получено {len(orders)}')
    orders = sorted(orders, key=lambda x: x['barcode'])
    barcodes = get_barcodes_with_full_info(orders)
    create_pdf_stickers_by_barcodes(barcodes)
    create_db_for_checking(barcodes)


def get_start_and_end_of_current_day():
    tz2 = pytz.timezone('Europe/Moscow')
    today = datetime.datetime.utcnow().date() - datetime.timedelta(1)
    start = datetime.datetime(today.year, today.month,
                              today.day, tzinfo=tz.tzutc()).astimezone(tz2)
    end = start + datetime.timedelta(1)
    return (start.replace(tzinfo=pytz.UTC).isoformat(),
            end.replace(tzinfo=pytz.UTC).isoformat())


def add_json_file_to_today_json(path_to_json_file):
    json_dir = create_all_today_path()['json_dir']
    filename = 'barcodes_%s.json' % datetime.datetime.now().strftime('%H%M')
    path_to_backup_file = os.path.join(json_dir, filename)
    shutil.copyfile(path_to_json_file, path_to_backup_file)


def add_results_file_to_today_backup(path_to_results_file):
    backup_dir = create_all_today_path()['backup_dir']
    filename = 'results_%s.pdf' % datetime.datetime.now().strftime('%H%M')
    path_to_backup_file = os.path.join(backup_dir, filename)
    shutil.copyfile(path_to_results_file, path_to_backup_file)


def get_list_of_relative_path_to_all_today_results():
    list_of_files = []
    backup_dir = create_all_today_path()['backup_dir']
    for root, directories, file in os.walk(backup_dir):
        for file in file:
            list_of_files.append(os.path.relpath((os.path.join(root, file))))
    return list_of_files


def get_list_of_relative_path_to_all_today_json():
    list_of_files = []
    json_dir = create_all_today_path()['json_dir']
    for root, directories, files in os.walk(json_dir):
        for file in files:
            if file != 'result_fbs.json':
                list_of_files.append(os.path.relpath(
                    (os.path.join(root, file))))
    return list_of_files


def get_list_of_relative_path_to_all_logs():
    list_of_files = []
    for root, directories, file in os.walk(log_dir):
        for file in file:
            list_of_files.append(os.path.relpath((os.path.join(root, file))))
    return list_of_files


def get_dict_of_unique_orders_and_article():
    list_of_json = get_list_of_relative_path_to_all_today_json()
    order_and_article_dict = {}
    for json_path in list_of_json:
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            for barcode in data.keys():
                orders = data[barcode]['orders']
                for order in orders:
                    order_and_article_dict[order] = data[barcode]['info']['article']
    return order_and_article_dict


def get_dict_of_unique_orders_and_nmid():
    list_of_json = get_list_of_relative_path_to_all_today_json()
    order_and_nmid_dict = {}
    for json_path in list_of_json:
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            for barcode in data.keys():
                orders = data[barcode]['orders']
                for order in orders:
                    order_and_nmid_dict[order] = data[barcode]['info']['nmid']
    return order_and_nmid_dict


def get_dict_of_unique_orders_and_nmid():
    list_of_json = get_list_of_relative_path_to_all_today_json()
    order_and_article_dict = {}
    for json_path in list_of_json:
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            for barcode in data.keys():
                orders = data[barcode]['orders']
                for order in orders:
                    order_and_article_dict[order] = data[barcode]['info']['nmid']
    return order_and_article_dict


def get_today_article_and_count():
    orders_and_article = get_dict_of_unique_orders_and_article()
    article_and_count = {}
    for order in orders_and_article.keys():
        article = orders_and_article[order]
        if not article in article_and_count:
            article_and_count[article] = 1
        else:
            article_and_count[article] += 1
    return article_and_count


def get_today_nmid_and_count():
    orders_and_article = get_dict_of_unique_orders_and_nmid()
    article_and_count = {}
    for order in orders_and_article.keys():
        article = orders_and_article[order]
        if not article in article_and_count:
            article_and_count[article] = 1
        else:
            article_and_count[article] += 1
    return article_and_count


def create_finall_table_of_day():
    logging.info('Получение артикулов и количества заказов для них')
    article_and_count = get_today_nmid_and_count()

    json_dir = create_all_today_path()['json_dir']
    path_for_saving_aricle_and_count_dict = os.path.join(
        json_dir, 'result_fbs.json')
    with open(path_for_saving_aricle_and_count_dict, 'w') as f:
        json.dump(article_and_count, f)
    book = openpyxl.Workbook()
    sheet = book.active
    sheet['A1'] = 'Артикул'
    sheet['A1'].border = THIN_BORDER
    sheet['B1'] = 'Количество'
    sheet['B1'].border = THIN_BORDER
    row = 2
    article_and_count = get_today_article_and_count()
    for article in article_and_count.keys():
        cell = sheet.cell(row=row, column=1)
        cell.value = article
        cell.border = THIN_BORDER

        cell = sheet.cell(row=row, column=2)
        cell.value = article_and_count[article]
        cell.border = THIN_BORDER
        row += 1
    cell = sheet.cell(row=row, column=1)
    cell.value = 'Сумма'
    cell.border = MEDIUM_BORDER

    cell = sheet.cell(row=row, column=2)
    cell.value = f'=SUM(B2:B{row-1})'
    cell.border = MEDIUM_BORDER
    today_path_with_name = create_all_today_path()['today_path_with_name']
    file_path = os.path.join(today_path_with_name, 'final_bd.xlsx')
    book.save(file_path)
    book.close()
    return file_path


def get_orderId_and_sticker_encoded(ids):
    url = 'https://suppliers-api.wildberries.ru/api/v2/orders/stickers'
    json_order_id = {
        "orderIds": ids
    }
    response = requests.post(
        url,
        json=json_order_id,
        headers=headers)
    order_and_sticker_encoded = {}
    data = response.json()['data']
    while len(data) == 0:
        response = requests.post(
            url,
            json=json_order_id,
            headers=headers)
        data = response.json()['data']

    for order in response.json()['data']:
        order_and_sticker_encoded[order['orderId']
                                  ] = order['sticker']['wbStickerEncoded']
    return order_and_sticker_encoded


def filter_orders_by_article(articles, count):
    orders = get_all_orders(status=0)
    barcodes = get_barcodes_with_full_info(orders)
    filtered_orders = []
    for barcode in barcodes:
        if (barcodes[barcode]['info']['article'] in articles) and (
                len(filtered_orders) <= count):
            filtered_orders += barcodes[barcode]['orders']
    return filtered_orders[:count]


def swap_token_by_name(name):
    global TOKEN
    cred = json.load(open('credentials.json', 'rb'))
    global NAME
    TOKEN = cred[name]['token']
    NAME = cred[name]['name']
    global headers
    headers = {
        'Authorization': TOKEN,
    }
    create_all_today_path()


def get_name():
    return NAME


def close_supplie(supplie_id: str):
    URL_FOR_CLOSING_SUPPLIE = f'https://suppliers-api.wildberries.ru/api/v2/supplies/{supplie_id}/close'
    response = requests.post(URL_FOR_CLOSING_SUPPLIE, headers=headers)
    if response.status_code == 204:
        return None
    return response.json()['errorText']


def add_orders_to_supplie(supplie_id, orders):
    try:
        order_ids = [order['orderId'] for order in orders]
        return add_orders_to_supplie_by_id(supplie_id, order_ids)
    except BaseException:
        return 'Что-то пошло не так'


def add_orders_to_supplie_by_id(supplie_id, orders_ids):
    data = {'orders': orders_ids}
    js = json.dumps(data)
    URL_FOR_ADD_ORDERS_TO_SUPPLIE = f'https://suppliers-api.wildberries.ru/api/v2/supplies/{supplie_id}'
    response = requests.put(
        URL_FOR_ADD_ORDERS_TO_SUPPLIE, headers=headers, data=js)

    if response.status_code != 200:
        return 200
    else:
        return response.json()['errorText']


def get_supplies(status: str = 'ACTIVE'):
    URL_FOR_GETTING_SUPPLIES = 'https://suppliers-api.wildberries.ru/api/v2/supplies'
    params = {'status': status}
    response = requests.get(
        URL_FOR_GETTING_SUPPLIES,
        params=params,
        headers=headers)
    return response.json()['supplies']


def create_new_supplie():
    URL_FOR_CREATING_SUPPLIE = 'https://suppliers-api.wildberries.ru/api/v2/supplies'
    response = requests.post(URL_FOR_CREATING_SUPPLIE, headers=headers)
    data = {
        'supplyId': '',
        'error': ''
    }
    error_dict = {
        409: 'У поставщика уже есть активная поставка',
        500: 'Ошибка WB'
    }
    if response.status_code == 201:
        data['supplyId'] = response.json()['supplyId']
        return data
    data['error'] = error_dict[response.status_code]
    return data


def get_data_svg_stick_of_supplie(supplie_id: str):
    data = {'data_for_svg': '', 'error': ''}
    URL_FOR_GETTING_STICK_OF_SUPPLIE = f'https://suppliers-api.wildberries.ru/api/v2/supplies/{supplie_id}/barcode'
    response = requests.get(
        URL_FOR_GETTING_STICK_OF_SUPPLIE,
        headers=headers,
        params={
            'type': 'svg'})
    if response.status_code == 200:
        data['data_for_svg'] = response.json()['file']
        return data
    data['error'] = response.json()['errorText']
    return data


def create_file_by_data(data: str, path_for_save: str):
    file_data = bytes(data, 'utf-8')
    with open(path_for_save, 'wb') as f:
        f.write(codecs.decode(file_data, 'base64'))


def create_stick_of_supplie_by_svg_and_name(
        path_to_svg: str, name: str, path_for_save: str):
    drawing = svg2rlg(path_to_svg)
    canvas = Canvas(path_for_save, pagesize=A4)
    canvas.setFont('FreeSans', 20)
    renderPDF.draw(drawing, canvas, 50 * mm, 190 * mm)
    canvas.drawString(
        70 * mm,
        170 * mm,
        f'{name} {datetime.datetime.now().strftime("%H:%M")}')
    canvas.save()


def create_stick_of_supplie(supplie):
    data = get_data_svg_stick_of_supplie(supplie_id=supplie)['data_for_svg']
    path_to_stick_svg = 'stick.svg'
    path_for_stick_pdf = 'stick.pdf'
    create_file_by_data(data, path_to_stick_svg)
    create_stick_of_supplie_by_svg_and_name(
        path_to_svg=path_to_stick_svg,
        name=NAME,
        path_for_save=path_for_stick_pdf)
    return path_for_stick_pdf

def create_db_by_file(file):
    orders = json.load(file)
    barcodes = get_barcodes_with_full_info(orders)
    
    create_db_for_checking(barcodes)

if __name__ == '__main__':
    # orders = get_all_orders(status=1)
    # print(orders)

    with open('barcodes.json', 'r', encoding='utf-8') as f:
        bar = json.load(f)
    create_db_for_checking(bar)
    # barcodes = get_barcodes_with_full_info(orders)
    # with open('barcodes.json', 'w', encoding='utf-8') as f:
    #     json.dump(barcodes, f, ensure_ascii=False)
    # add_json_file_to_today_json('barcodes.json')
    # create_pdf_stickers_by_barcodes(barcodes)
    # return (len(orders), barcodes)
