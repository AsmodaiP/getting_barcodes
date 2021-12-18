import datetime
import json
from sys import path
from typing import Dict, List
from openpyxl.styles.fills import PatternFill
import pytz
import requests
import shutil
import codecs
import os
from dotenv import load_dotenv
from PyPDF2 import PdfFileMerger
from reportlab.pdfgen.canvas import Canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.units import inch
import logging
from logging.handlers import RotatingFileHandler
import openpyxl
from openpyxl.formula.translate import Translator
from requests.models import ReadTimeoutError
from dateutil import tz
from openpyxl.styles.borders import Border, Side
import sys
from openpyxl.styles import Protection, Font, Fill
from openpyxl.formatting import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import CellIsRule
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from svglib.svglib import svg2rlg
from reportlab.graphics import renderPDF
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
import create_stickers_and_db

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
MEDIUM_BORDER = Border(left=Side(style='medium'),
                       right=Side(style='medium'),
                       top=Side(style='medium'),
                       bottom=Side(style='medium'))
THIN_BORDER = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

cred = json.load(open('credentials.json', 'rb'))
TOKEN = cred['Савельева']['token']
pdfmetrics.registerFont(TTFont('FreeSans', 'fonts/FreeSans.ttf'))

# token='eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJhY2Nlc3NJRCI6IjdmYTY4YTgyLTQ4ODUtNDAyZi04YzAwLTRmYTUxNTE1YzBiZSJ9.5u_vveflVF8Nl8tl8wAxbN1L3F_BA2EWbUoYyLSnXvo'

def get_supplies(token:str, status:str='ACTIVE'):
    headers = {
        'Authorization': token,
    }
    URL_FOR_GETTING_SUPPLIES = 'https://suppliers-api.wildberries.ru/api/v2/supplies'
    params = {'status': status}
    response = requests.get(URL_FOR_GETTING_SUPPLIES, params=params, headers=headers)
    return response.json()['supplies']

def create_new_supplie(token:str):
    URL_FOR_CREATING_SUPPLIE = 'https://suppliers-api.wildberries.ru/api/v2/supplies'
    headers = {
        'Authorization': token,
    }
    response = requests.post(URL_FOR_CREATING_SUPPLIE,headers=headers)
    data = {
        'supplyId': '',
        'error':''
    }
    error_dict={
        409: 'У поставщика уже есть активная поставка',
        500: 'Ошибка WB'
    }
    if response.status_code == 201:
        data['supplyId']=response.json()['supplyId']
        return response.json()['supplyId']
    data['error'] = error_dict[response.status_code]
    return data

def add_orders_to_supplie(token: str, supplie_id: str, orders: List):
    order_ids = [order['orderId'] for order in orders]
    add_orders_to_supplie_by_id(token, supplie_id, order_ids)

def add_orders_to_supplie_by_id(token:str, supplie_id:str, orders_ids:List[str]):
    headers = {
        'Authorization': token,
    }
    data = {'orders':orders_ids}
    js = json.dumps(data)
    URL_FOR_ADD_ORDERS_TO_SUPPLIE = f'https://suppliers-api.wildberries.ru/api/v2/supplies/{supplie_id}'
    response = requests.put(URL_FOR_ADD_ORDERS_TO_SUPPLIE, headers=headers, data=js)
    # print(response.content.decode('utf-8'))
    if response.status_code != 200:
        return 200
    else:
        return response.json()['errorText']


def close_supplie(token:str, supplie_id:str):
    headers = {
        'Authorization': token,
    }
    URL_FOR_CLOSING_SUPPLIE = f'https://suppliers-api.wildberries.ru/api/v2/supplies/{supplie_id}/close'
    response = requests.post(URL_FOR_CLOSING_SUPPLIE, headers=headers)
    if response.status_code == 200:
        return None
    return response.json()['errorText']

def get_data_svg_stick_of_supplie(token:str, supplie_id:str):
    headers = {
        'Authorization': token,
    }
    data = {'data_for_creating_pdf': '', 'error': ''}
    URL_FOR_GETTING_STICK_OF_SUPPLIE =  f'https://suppliers-api.wildberries.ru/api/v2/supplies/{supplie_id}/barcode'
    response = requests.get(URL_FOR_GETTING_STICK_OF_SUPPLIE, headers=headers, params={'type': 'svg'})
    if response.status_code == 200:
        data['data_for_creating_pdf'] = response.json()['file']
        return data
    data['error']=response.json()['errorText']
    return data


def create_file_by_data(data:str, path_for_save:str):
    file_data = bytes(data, 'utf-8')
    with open(path_for_save, 'wb') as f:
        f.write(codecs.decode(file_data, 'base64'))

def create_stick_of_supplie_by_svg_and_name(path_to_svg:str, name:str, path_for_save:str):
    drawing = svg2rlg(path_to_svg)
    canvas = Canvas(path_for_save, pagesize=A4)
    canvas.setFont('FreeSans', 20)
    renderPDF.draw(drawing, canvas,50*mm,190*mm)
    canvas.drawString(70*mm, 170*mm, name)
    canvas.save()


def get_suplies_orders(token:str, supplie_id:str):
    URL_FOR_ORDERS_OF_SUPPLIE =  f'https://suppliers-api.wildberries.ru/api/v2/supplies/{supplie_id}/orders'
    headers = {
        'Authorization': token,
    }
    response = requests.get(URL_FOR_ORDERS_OF_SUPPLIE, headers=headers)
    data = {'orders': '', 'error':  ''}
    if response.status_code == 200:
        data['orders'] = response.json()['orders']
        return data
    data['error']= response.json()['errorText']
    return data


def get_barcodes_with_orders_and_chartId(token, orders):
    barcodes_and_ids = {}
    logging.info('Сортировка информации по баркодам')
    for order in orders:
        barcode = order['barcodes'][0]
        id = int(order['orderId'])
        chrt_id = order['chrtId']
        if barcode not in barcodes_and_ids.keys():
            barcodes_and_ids[barcode] = {'orders': [id], 'chrtId': chrt_id}
        else:
            barcodes_and_ids[barcode]['orders'] += [id]

    logging.info(f'Получено {len(barcodes_and_ids)} баркодов')
    return barcodes_and_ids


def add_information_about_barcodes_and_len(token,barcodes):
    for barcode in barcodes.keys():
        barcodes[barcode]['info'] = getting_information_about_barcode_by_chartId(token,barcodes[barcode]['chrtId'])
        barcodes[barcode]['info']['count'] = len(barcodes[barcode]['orders'])
    return barcodes

def get_card_by_chrtId(token, chrtId):
    headers = {
        'Authorization': token,
    }
    # print(token)
    url = 'https://suppliers-api.wildberries.ru/card/list'
    json_for_request = {
        "id": 1,
        "jsonrpc": "2.0",
        "params": {
            "filter": {
                "find": [
                    {
                        "column": "nomenclatures.variations.chrtId",
                        "search": chrtId
                    }
                ],
                "order": {
                    "column": "string",
                    "order": "string"
                }
            }
        }
    }
    response = requests.post(url=url, headers=headers, json=json_for_request)
    # print(response.content.decode('utf-8'))
    card = response.json()['result']['cards'][0]
    return card

def get_data_nomenclature_from_card_by_chrtId(card, chrtId):
    all_nomenclatures = card['nomenclatures']
    for nomenclature in all_nomenclatures:
        for diffetent_types in nomenclature['variations']:
            vendorCode = nomenclature['vendorCode']
            data_about_nomenclature = diffetent_types
            for field in data_about_nomenclature:
                if data_about_nomenclature[field] == chrtId:
                    return(data_about_nomenclature, vendorCode, nomenclature)

def getting_information_about_barcode_by_chartId(token, chrtId):
    good = get_card_by_chrtId(token, chrtId)
    
    supplierVendorCode = good['supplierVendorCode']
    name = ''
    nomenclature_data, article, nomenclature = get_data_nomenclature_from_card_by_chrtId(
        good, int(chrtId))
    article =  supplierVendorCode + article 
    size = ''
    color = ''
    extra_colors = ''
    addin = nomenclature['addin']
    nmid=nomenclature['nmId']
    for type_and_params in addin:
        if type_and_params['type'] in ('Цвет', 'Основной цвет'):
            color = type_and_params['params'][0]['value']
        if type_and_params['type'] == 'Доп. цвета':
            for extra_color in type_and_params['params']:
                extra_colors += extra_color['value']+' '
    if 'Доп. цвета' in nomenclature.keys():
        extra_colors = nomenclature['Доп. цвета']
    for type_and_params in good['addin']:
        if type_and_params['type'] == 'Наименование':
            name = type_and_params['params'][0]['value']
    for data in nomenclature_data['addin']:
        if data['type'] == 'Размер':
            size = data['params'][0]['value']
        if data['type'] == 'Доп. цвета':
            extra_colors = data['params'][0]['value']
    info = {
        'name': name,
        'article': article,
        'chrtId': chrtId,
        'size': size,
        'color': color,
        'extra_colors': extra_colors,
        'nmId': nmid
    }
    return info

def sorted_barcodes_by_count_of_orders(barcodes):
    sorted_tuples = sorted(barcodes.items(), key=lambda x: len(
        x[1]['orders']), reverse=True)
    sorted_dict = {k: v for k, v in sorted_tuples}
    return sorted_dict

def get_barcodes_with_full_info(token, orders):
    barcodes = get_barcodes_with_orders_and_chartId(token, orders)
    barcodes = add_information_about_barcodes_and_len(token, barcodes)
    barcodes = sorted_barcodes_by_count_of_orders(barcodes)
    return barcodes


def add_json_file_to_today_json(name, path_to_json_file):
    json_dir = create_all_today_path(name)['json_dir']
    filename = 'barcodes_%s.json' % datetime.datetime.now().strftime('%H%M')
    path_to_backup_file = os.path.join(json_dir, filename)
    shutil.copyfile(path_to_json_file, path_to_backup_file)

def create_path_if_not_exist(path):
    if not os.path.exists(path):
        os.mkdir(path)

def create_all_today_path(name):
    pdf_path = os.path.join(BASE_DIR, 'pdf/')
    create_path_if_not_exist(pdf_path)
    today_prefix_path = os.path.join(
        pdf_path, datetime.datetime.today().strftime('%Y_%m_%d'))
    create_path_if_not_exist(today_prefix_path)
    today_path_with_name = os.path.join(today_prefix_path, name)
    create_path_if_not_exist(today_path_with_name)
    backup_dir = os.path.join(today_path_with_name, 'results/')
    json_dir = os.path.join(today_path_with_name, 'json/')
    create_path_if_not_exist(backup_dir)
    create_path_if_not_exist(json_dir)
    return {
        'backup_dir': backup_dir,
        'json_dir': json_dir,
        'today_prefix_path': today_prefix_path,
        'today_path_with_name':today_path_with_name
    }


def edit_blank_pdf(barcode_info, name):

    canvas = Canvas('pdf/blank1.pdf', pagesize=(1.6 * inch, 1.2 * inch))
    font_size = 9
    canvas.setFont('FreeSans', font_size)
    n = 0
    blanks_number = 1
    slice = 23
    name_of_host = name
    name = barcode_info['name']
    size = barcode_info['size']
    article = f'article = {barcode_info["article"]}'
    color = barcode_info['color']
    extra_colors = barcode_info['extra_colors']
    count = 'Количество = ' + str(barcode_info['count'])
    chrtId = f'chrtId = {barcode_info["chrtId"]}'
    blanks = ['pdf/blank1.pdf']
    for params in (count, name_of_host, name, size, color, extra_colors, article, chrtId):
        a = str(params)
        while len(a) > 0:
            if n > 8:
                blanks_number += 1
                canvas.save()
                canvas = Canvas(
                    f'pdf/blank{blanks_number}.pdf', pagesize=(1.6 * inch, 1.2 * inch))
                canvas.setFont('FreeSans', font_size)
                blanks += [f'pdf/blank{blanks_number}.pdf']
                n = 0
            canvas.drawString(2, 75 - n * font_size,
                              a[:slice], charSpace=0)
            a = a[slice:]
            n += 1
    canvas.save()
    merg = PdfFileMerger()
    for blank in blanks:
        merg.append(blank)
    merg.write('pdf/blank.pdf')
    merg.close()
    logging.info(
        f'Получена информация: наименование "{barcode_info["name"]}" '
        f' {count}, '
        f'артикул = {article}, '
        f'chrtId = {chrtId} '
        f'{"Size "+size if size != "" else ""} '
        f'{"color "+ color if color != "" else ""} '
        f'{"Доп цвета: "+extra_colors if extra_colors != "" else ""} ')

def create_and_merge_pdf_by_barcodes_and_ids(token, name, barcodes_and_ids):
    headers = {
        'Authorization': token,
    }
    logging.info('Создание pdf для баркодов')
    results_files = []
    url_for_getting_stikers = 'https://suppliers-api.wildberries.ru/api/v2/orders/stickers/pdf'
    # print(barcodes_and_ids.keys())
    for barcode in barcodes_and_ids.keys():
        
        edit_blank_pdf(barcodes_and_ids[barcode]['info'], name)
        pdfs = ['pdf/blank.pdf']
        merger = PdfFileMerger()
        orders = barcodes_and_ids[barcode]['orders']
        json_orders_id = {
            "orderIds": orders
        }
        response = requests.post(
                url_for_getting_stikers,
                json=json_orders_id,
                headers=headers)
        data_for_pdf = response.json()['data']['file']
        file_data = bytes(data_for_pdf, 'utf-8')
        today_path_with_name = create_all_today_path(name)['today_path_with_name']
        path = os.path.join(today_path_with_name, f'{barcode}.pdf')
        with open(path, 'wb') as f:
            f.write(codecs.decode(file_data, 'base64'))
        pdfs += [path]
        for pdf in pdfs:
            merger.append(pdf)
        path_for_result_of_barcode = os.path.join(
            today_path_with_name, f'result_{barcode}.pdf')
        merger.write(path_for_result_of_barcode)
        results_files.append(path_for_result_of_barcode)
        merger.close()
        logging.info(f'Создано pdf для баркода {barcode}')
    return results_files

def add_results_file_to_today_backup(name, path_to_results_file):
    backup_dir = create_all_today_path(name)['backup_dir']
    filename = 'results_%s.pdf' % datetime.datetime.now().strftime('%H%M')
    path_to_backup_file = os.path.join(backup_dir, filename)
    shutil.copyfile(path_to_results_file, path_to_backup_file)

def create_pdf_stickers_by_barcodes(token,name, barcodes_and_ids):
    results_files = []
    results_files = create_and_merge_pdf_by_barcodes_and_ids(token, name, barcodes_and_ids)
    merger = PdfFileMerger()
    logging.info('Объединение pdf файлов в results.pdf')
    for result in results_files:
        merger.append(result)
    merger.write('results.pdf')
    merger.close()
    add_results_file_to_today_backup(name,'results.pdf')

def create_stickers_by_supplie_id(token,name, supplie_id):
    orders = get_suplies_orders(token, supplie_id)['orders']
    if len(orders) == 0:
        return (0,0)
    barcodes = get_barcodes_with_full_info(token,orders)
    
    with open('barcodes.json','w', encoding='utf-8') as f:
        json.dump(barcodes, f,ensure_ascii=False)
    add_json_file_to_today_json(name,'barcodes.json')
    create_pdf_stickers_by_barcodes(token, name, barcodes)
    return (len(orders), barcodes)

def get_now_time():
    d = datetime.datetime.utcnow()
    d_with_timezone = d.replace(tzinfo=pytz.UTC)
    return(d_with_timezone.isoformat())

def get_all_orders(token, status=0, date_end=get_now_time(), date_start='2021-11-06T00:47:17.528082+00:00'):
    headers = {
        'Authorization': token,
    }
    URL_FOR_GETTING_ORDERS = 'https://suppliers-api.wildberries.ru/api/v2/orders'
    logging.info(f'Получение всех заказов со статусом {status}')
    date_end=get_now_time()
    orders = []

    params = {
        'date_end': date_end,
        'date_start': date_start,
        'status': status,
        'take': 1000,
        'skip': 0
    }
    response = requests.get(
        URL_FOR_GETTING_ORDERS,
        headers=headers,
        params=params)
    try:
        orders_from_current_response = response.json()['orders']
    except KeyError as e:
        orders_from_current_response = []
        logging.error(e, exc_info=True)
    orders += orders_from_current_response
    while orders_from_current_response != []:
        params['skip'] += len(orders_from_current_response)
        response = requests.get(
            URL_FOR_GETTING_ORDERS,
            headers=headers,
            params=params)
        orders_from_current_response = response.json()['orders']
        orders += orders_from_current_response
        logging.info(f'{len(orders)}')
    logging.info(f'Получено {len(orders)}')
    return orders

def get_StickerEncoded_by_orderId(token, id):
    headers = {
        'Authorization': token,
    }
    url = 'https://suppliers-api.wildberries.ru/api/v2/orders/stickers'
    json_order_id = {
        "orderIds": [int(id)]
    }
    response = requests.post(
        url,
        json=json_order_id,
        headers=headers)
    return response.json()['data'][0]['sticker']['wbStickerEncoded']

def get_orderId_and_sticker_encoded(token, ids):
    headers = {
        'Authorization': token,
    }
    url = 'https://suppliers-api.wildberries.ru/api/v2/orders/stickers'
    json_order_id = {
        "orderIds": ids
    }
    response = requests.post(
        url,
        json=json_order_id,
        headers=headers)
    order_and_sticker_encoded = {}
    for order in response.json()['data']:
        order_and_sticker_encoded[order['orderId']
                                  ] = order['sticker']['wbStickerEncoded']
    return order_and_sticker_encoded
def create_db_for_checking(token,barcodes):
    barcodes_and_stickers = {}
    logging.info('Получение расшифрованных стикеров')
    article_counts = {}
    for barcode in barcodes.keys():
        logging.info(f'Получение расшифрованных стикеров для {barcode}')
        barcodes_and_stickers[barcode] = {}
        arcticle = barcodes[barcode]['info']['article']
        name = barcodes[barcode]['info']['name']
        size = barcodes[barcode]['info']['size']
        color = barcodes[barcode]['info']['color']
        orders_and_sticker_encoded = get_orderId_and_sticker_encoded(token,
            barcodes[barcode]['orders'])
        for order in barcodes[barcode]['orders']:
            sticker_encoded = orders_and_sticker_encoded[order]
            barcodes_and_stickers[barcode][order] = {
                'sticker_encoded': sticker_encoded,
                'article': arcticle,
                'name': name[:40],
                'size': size,
                'color': color
            }
            if not arcticle in article_counts:
                article_counts[arcticle] = 1
            else:
                article_counts[arcticle] += 1

    book = openpyxl.Workbook()
    sheet = book.active
    book.create_sheet("Итог")
    book.create_sheet('Проверка')
    sheet['A1'] = 'Номер заказа'
    sheet['B1'] = 'Артикул'
    sheet['C1'] = 'Наименование'
    sheet['D1'] = 'Баркод'
    sheet['E1'] = 'Stick'
    sheet.protection.sheet=True
    sheet.protection.set_password('osdfjl2')
    sheet.protection.enable()

    row = 2
    logging.info('Формирование xlsx файла')
    for barcode in barcodes_and_stickers.keys():
        barcode_info = barcodes_and_stickers[barcode]
        for order in barcode_info.keys():
            info = barcode_info[order]
            article = info['article']
            sticker_encoded = info['sticker_encoded']
            name = info['name']
            sheet[row][0].value = str(order)
            sheet[row][1].value = str(article)
            sheet[row][2].value = name
            sheet[row][3].value = str(barcode)
            sheet[row][4].value = sticker_encoded
            sheet[f'K{row}'] = f'=IF(ISERROR(MATCH(J{row+1},L{row},0)),"",TRUE)'
            sheet[f'L{row}'] = f'=IF(ISERROR(INDEX(D:D,MATCH(J{row},E:E,0),1)),"",INDEX(D:D,MATCH(J{row},E:E,0),1))'
            row += 1

    # book.create_sheet("Итог")
    book.active = 1                                                     
    sheet = book.active
    sheet['A1'] = 'Артикул'
    sheet['A1'].border = THIN_BORDER
    sheet['B1'] = 'Количество'
    sheet['B1'].border = THIN_BORDER
    sheet['C1'] = 'Собрано'
    sheet['B1'].border = THIN_BORDER
    row = 2
    sheet.protection.sheet=True
    sheet.protection.set_password('osdfjl2')
    sheet.protection.enable()
    for article in article_counts.keys():
        cell = sheet.cell(row=row, column=1)
        cell.value = article
        cell.border = THIN_BORDER

        cell = sheet.cell(row=row, column=2)
        cell.value = article_counts[article]
        cell.border = THIN_BORDER

        cell = sheet.cell(row=row, column=3)
        cell.value = f'=COUNTIF(Проверка!D:D,A{row})'
        cell.border = THIN_BORDER
        row += 1
    cell = sheet.cell(row=row, column=1)
    cell.value = 'Сумма'
    cell.border = MEDIUM_BORDER

    cell = sheet.cell(row=row, column=2)
    cell.value = f'=SUM(B2:B{row-1})'
    cell.border = MEDIUM_BORDER
    cell = sheet.cell(row=row, column=3)
    cell.value = f'=SUM(C2:C{row-1})'
    cell.border = THIN_BORDER
    
    green_fill=PatternFill(bgColor="a0db8e")
    sheet.conditional_formatting.add('C1:C3000',FormulaRule(formula=['IF(AND(C1=B1, C1<>""),True,False)'], stopIfTrue=True, fill=green_fill))
 
    book.active = 2
    sheet = book.active
    red_fill = PatternFill(bgColor="FFC7CE")

    sheet.conditional_formatting.add('B1:B3000',FormulaRule(formula=['IF(B1="Ошибка",True,False)'], stopIfTrue=True, fill=red_fill))

    row = 2
    sheet['A1'] = 'Данные'
    sheet['B1'] = 'Результат'
    sheet['D1'] = 'Артикул'
    sheet.protection.sheet=True
    sheet.protection.enable()
    for barcode in barcodes_and_stickers.keys():
        barcode_info = barcodes_and_stickers[barcode]
        for order in barcode_info.keys():
            for i in range(2):
                if row % 2 == 0: 
                    # =IF(AND(ISERROR(MATCH(A3,C2,0)),A2<>"",A3<>""),"Ошибка","")
                    sheet[f'B{row}'] = f'=IF(AND(ISERROR(MATCH(A{row+1},C{row},0)),A{row}<>"",A{row+1}<>""),"Ошибка","")'
                    sheet[f'C{row}'] = f'=IF(ISERROR(INDEX(Sheet!D:D,MATCH(A{row},Sheet!E:E,0),1)),"",INDEX(Sheet!D:D,MATCH(A{row},Sheet!E:E,0),1))'
                else:
                    sheet[f'D{row}'] = f'=IF(AND(B{row-1}<>"Ошибка", A{row-1}<>""),INDEX(Sheet!B:B, MATCH(A{row},Sheet!D:D,0),1), "")'
                sheet[f'A{row}'].number_format = '@'
                sheet[f'A{row}'].protection = Protection(locked=False)
                row += 1
    for column in ['C', 'D']:
        for cell in sheet[column]:
            cell.font= Font(color='Ffffffff')
    book.save('db.xlsx')
    book.close()

def create_stickers_by_id(token, name, ids):
    headers = {
        'Authorization': token,
    }
    URL_FOR_GETTING_ORDERS = 'https://suppliers-api.wildberries.ru/api/v2/orders'
    date_end = get_now_time()
    orders = []
    for id in ids:
        date_start = '2021-11-06T00:47:17.528082+00:00'

        params = {
            'date_end': date_end,
            'date_start': date_start,
            'status': 2,
            'take': 100,
            'skip': 0,
            'id': id
        }
        response = requests.get(
            URL_FOR_GETTING_ORDERS,
            headers=headers,
            params=params)
        try:
            orders_from_current_response = response.json()['orders']
        except KeyError as e:
            logging.error(e, exc_info=True)
        orders += orders_from_current_response
        logging.info(f'Получено {len(orders)}')
    orders = sorted(orders, key=lambda x: x['barcode'])
    barcodes = get_barcodes_with_full_info(token,orders)
    with open('barcodes.json','w', encoding='utf-8') as f:
        json.dump(barcodes, f,ensure_ascii=False)
    add_json_file_to_today_json(name, 'barcodes.json')
    create_pdf_stickers_by_barcodes(token,name,barcodes,)
    create_db_for_checking(token,barcodes)



if __name__ == '__main__':
    pass

    # print(get_supplies(token))
    # orders = get_all_orders(token, status=1)

    # print(add_orders_to_supplie(token, supplie_id='WB-GI-4858239', orders=orders))
    # orders = get_suplies_orders(token, 'WB-GI-4858872')['orders']

    # print(close_supplie(token=token,supplie_id='WB-GI-4860760'))
    # print(len(orders))
